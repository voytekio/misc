import pdb
import sys

import openpyxl

class Xcel():
    def __init__(self, file_loc):
        self.file_loc = file_loc
        #self.wb = openpyxl.Workbook(iso_dates=True)
        self.wb = openpyxl.load_workbook(self.file_loc)

    def save_xls(self, suffix='1'):
        file_split = self.file_loc.split('.')
        new_file_name = '{}_{}.xlsx'.format(file_split[0], suffix)
        print('Saving xls file: {}'.format(new_file_name))
        self.wb.save(new_file_name)

    def resolve_range(self, range_value):
        #pdb.set_trace()
        ret = {}
        starts = range_value.split(':')[0]
        ends = range_value.split(':')[1]
        for num, char in enumerate(starts):
            if char.isdigit():
                index = num
                break
        column_as_int = ord(starts[0].upper()) - 64 # we take the first char only as in 'd' in 'D123'
        ret['column_start'] = int(column_as_int)
        ret['row_start'] = int(starts[index:])

        for num, char in enumerate(ends):
            if char.isdigit():
                index = num
                break
        column_as_int = ord(ends[0].upper()) - 64 # we take the first char only as in 'd' in 'D123'
        ret['column_end'] = int(column_as_int)
        ret['row_end'] = int(ends[index:])
        return ret
                
    def is_custom_date(self, cell_value):
        if isinstance(cell_value, str) and cell_value.count('.') == 2 and len(cell_value) > 2 and cell_value[2] == '.':
            print('{} is a Custom_date'.format(cell_value))
            return True
        print('{} is a not a date object'.format(cell_value))
        return False

    def replace_dates(self, sheet_name, range_value):
        res = self.resolve_range(range_value)
        sheet = self.wb.get_sheet_by_name(sheet_name)
        for row_ref in range(res.get('row_start'), res.get('row_end')+1):
            for column_ref in range(res.get('column_start'), res.get('column_end')+1):
                cell = sheet.cell(column=column_ref, row=row_ref)
                if self.is_custom_date(cell.value):
                    cell.value = cell.value.replace('.','/')

    def print_cells(self, sheet_name, range_value):
        res = self.resolve_range(range_value)
        sheet = self.wb.get_sheet_by_name(sheet_name)
        #pdb.set_trace()
        for row_ref in range(res.get('row_start'), res.get('row_end')+1):
            for column_ref in range(res.get('column_start'), res.get('column_end')+1):
                cell = sheet.cell(column=column_ref, row=row_ref)
                print(cell.value, end="\t")
            print('\n')
        return True

    def replace_commas(self, sheet_name, range_value):
        res = self.resolve_range(range_value)
        sheet = self.wb.get_sheet_by_name(sheet_name)
        for row_ref in range(res.get('row_start'), res.get('row_end')+1):
            for column_ref in range(res.get('column_start'), res.get('column_end')+1):
                cell = sheet.cell(column=column_ref, row=row_ref)
                cell.value = float(cell.value.replace(',','.')) if isinstance(cell.value, str) else cell.value


def main():
    xls_location = sys.argv[1]
    x1 = Xcel(xls_location)
    #pdb.set_trace()
    x1.replace_commas('play', 'd1:d113')
    x1.replace_dates('play', 'a1:a113')
    x1.print_cells('play', 'a1:d113')
    x1.save_xls()

if __name__ == '__main__':
    main()
